import numpy as np
import pandas as pd
from openpyxl import Workbook,load_workbook
import pyinputplus as pyip
from pandas.io.formats.excel import ExcelFormatter
from pandas.io.excel._base import ExcelWriter
import os

def generate_class_scores(num_classes:int,file_name:str) -> None:
    '''
    param:num_classes,是班級數量
    param:file_name,自訂檔名,使用者要輸入副檔名xlsx

    1.建立新檔
    2.已經有新檔則覆蓋原來的檔案
    '''
    # 創建一個空的 Excel 工作簿
    wb = Workbook()
    
    # 將 Excel 工作簿保存為一個文件
    excel_path = file_name

    with ExcelWriter(excel_path, engine='openpyxl') as writer:
        for i in range(1, num_classes + 1):
            # 生成隨機分數
            scores = np.random.randint(0, 100, size=(3, 5))
            scores_df = pd.DataFrame(scores,
                                     columns=['國文', '英文', '數學', '地理', '歷史'],
                                     index=range(1, 4))

            # 計算總分和平均
            sum_values = scores_df.sum(axis=1)
            mean_values = scores_df.mean(axis=1)
            scores_df['總分'] = sum_values
            scores_df['平均'] = mean_values

            # 格式化和高亮
            pdstyle = scores_df.style \
                .highlight_between(left=0, right=59, props='color:white; background-color:#CB1B45;') \
                .highlight_max(axis=1, subset=scores_df.columns[:5], props='color:white; background-color:#2B5F75')

            # 將分數表寫入 Excel 工作簿的一個 sheet 中
            sheet_name = f'3年{i}班'
            pdstyle.to_excel(writer, sheet_name=sheet_name, index=False, engine='openpyxl')

def generate_and_add_class_scores(add_classes:int, file_name:str) -> None:
    # 載入現有的 Excel 工作簿
    wb = load_workbook(file_name)

    # 獲取現有工作簿的工作表數量
    num_sheets = len(wb.sheetnames)

    # 班級編號
    j = num_sheets + 1

    # 計算要添加的班級總數
    max_classes = num_sheets + add_classes

    for i in range(j, max_classes + 1):
        # 生成隨機分數
        scores = np.random.randint(0, 100, size=(3, 5))
        scores_df = pd.DataFrame(scores,
                                 columns=['國文', '英文', '數學', '地理', '歷史'],
                                 index=range(1, 4))

        # 計算總分和平均
        sum_values = scores_df.sum(axis=1)
        mean_values = scores_df.mean(axis=1)
        scores_df['總分'] = sum_values
        scores_df['平均'] = mean_values

        # 格式化和高亮
        pdstyle = scores_df.style \
            .format(precision=2) \
            .highlight_between(left=0, right=59, props='color:white; background-color:#CB1B45;') \
            .highlight_max(axis=1, subset=scores_df.columns[:5], props='color:white; background-color:#2B5F75')

        # 將分數表寫入 Excel 工作簿的一個新 sheet 中
        sheet_name = f'3年{i}班'
        with ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            pdstyle.to_excel(writer, sheet_name=sheet_name, index=False)
# 輸入班級數量
if __name__ == "__main__":
    num_classes = pyip.inputInt("請輸入班級數量(1~5):",min=1,max=5)     
    file_name = pyip.inputFilename("請輸入存檔名稱(xlsx):")
    if os.path.exists(file_name):
        is_covert = pyip.inputYesNo("請問需要覆蓋嗎(y,n)?",default="yes",blank=True)
        if is_covert == "no":
            generate_class_scores(num_classes,file_name)
            print(f"{file_name}已經建立")
        
    generate_class_scores(num_classes,file_name)
    print(f"{file_name}已經建立")